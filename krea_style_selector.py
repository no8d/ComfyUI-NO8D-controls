from __future__ import annotations

import json
import csv
import io
import importlib
import asyncio
import os
import re
import random
import unicodedata
import tempfile
import zipfile
from pathlib import Path
from io import BytesIO
from uuid import uuid4
from urllib.parse import parse_qs, quote, urlparse
from urllib.request import urlopen

try:
    import folder_paths
except ImportError:
    folder_paths = None

try:
    from aiohttp import web
    from server import PromptServer
except ImportError:
    web = None
    PromptServer = None


_ROOT = Path(__file__).resolve().parent
_BUILTIN_LIBRARY_DIR = _ROOT / "data" / "krea_style_libraries"
_API_PREFIX = "/no8d/krea-style-selector"
_CATEGORIES = ("写实摄影", "动漫插图", "手绘艺术", "数字艺术")
_DEFAULT_USER_CATEGORY = "全部"
_ALLOWED_IMAGE_TYPES = {"image/png": ".png", "image/jpeg": ".jpg", "image/webp": ".webp"}
_MAX_PREVIEW_BYTES = 15 * 1024 * 1024
_VIRTUAL_LIBRARIES = ("收藏夹", "历史记录")
_OPENPYXL_REQUIRED_MESSAGE = (
    "XLSX import and export require openpyxl, but it is not installed in ComfyUI's Python environment. "
    "Install this node's requirements.txt with ComfyUI Manager or ComfyUI's Python, then restart ComfyUI."
)


def _require_openpyxl():
    try:
        return importlib.import_module("openpyxl")
    except ImportError as error:
        raise RuntimeError(_OPENPYXL_REQUIRED_MESSAGE) from error


def _user_root() -> Path:
    if folder_paths is not None:
        getter = getattr(folder_paths, "get_user_directory", None)
        if callable(getter):
            return Path(getter()) / "no8d" / "krea_style_selector"
    override = os.environ.get("NO8D_KREA_USER_DIR")
    return Path(override) if override else _ROOT / "user_data" / "krea_style_selector"


def _read_user_payload() -> dict:
    root = _user_root()
    wildcard_path = root / "wildcards" / "custom.txt"
    metadata_path = root / "metadata.json"
    legacy_path = root / "styles.json"
    if not wildcard_path.is_file() and legacy_path.is_file():
        try:
            with legacy_path.open("r", encoding="utf-8") as handle:
                legacy = json.load(handle)
            if isinstance(legacy, dict) and isinstance(legacy.get("styles"), list):
                _write_user_payload({"version": 1, "styles": legacy["styles"]})
        except (OSError, json.JSONDecodeError):
            pass
    wildcard_dir = root / "wildcards"
    if not wildcard_dir.is_dir():
        return {"version": 2, "styles": []}
    try:
        metadata = {}
        metadata_version = 3
        if metadata_path.is_file():
            with metadata_path.open("r", encoding="utf-8") as handle:
                raw_metadata = json.load(handle)
            metadata = raw_metadata.get("entries", {}) if isinstance(raw_metadata, dict) else {}
            metadata_version = int(raw_metadata.get("version", 2)) if isinstance(raw_metadata, dict) else 2
        styles = []
        for path in sorted(wildcard_dir.rglob("*.txt")):
            library = path.relative_to(wildcard_dir).with_suffix("").as_posix()
            with path.open("r", encoding="utf-8-sig", errors="replace") as handle:
                for raw_line in handle:
                    line = raw_line.strip()
                    if not line:
                        continue
                    name, separator, prompt = line.partition(":")
                    if not separator or not name.strip() or not prompt.strip():
                        continue
                    name = name.strip()
                    key = f"{library}/{name}"
                    candidate = metadata.get(key, metadata.get(name, {}))
                    meta = candidate if isinstance(candidate, dict) else {}
                    category = str(meta.get("category") or _DEFAULT_USER_CATEGORY)
                    if metadata_version < 3 and category in _CATEGORIES:
                        category = _DEFAULT_USER_CATEGORY
                    styles.append({
                        "id": str(meta.get("id") or uuid4().hex),
                        "name": name,
                        "name_zh": str(meta.get("name_zh") or name),
                        "category": category,
                        "prompt": prompt.strip(),
                        "preview": str(meta.get("preview") or ""),
                        "library": library,
                    })
    except (OSError, json.JSONDecodeError):
        return {"version": 3, "styles": []}
    return {"version": 3, "styles": styles}


def _write_user_payload(payload: dict) -> None:
    root = _user_root()
    wildcard_dir = root / "wildcards"
    wildcard_dir.mkdir(parents=True, exist_ok=True)
    styles = payload.get("styles", [])
    libraries = {}
    for item in styles:
        library = _safe_library(item.get("library") or "custom")
        libraries.setdefault(library, []).append(item)
    metadata = {
        "version": 3,
        "entries": {
            f'{_safe_library(item.get("library") or "custom")}/{str(item["name"]).strip()}': {
                "id": item.get("id") or uuid4().hex,
                "name_zh": item.get("name_zh") or item["name"],
                "category": item.get("category") or _DEFAULT_USER_CATEGORY,
                "preview": item.get("preview") or "",
            }
            for item in styles
            if str(item.get("name", "")).strip()
        },
    }
    targets = []
    for library, entries in libraries.items():
        target = wildcard_dir / f"{library}.txt"
        target.parent.mkdir(parents=True, exist_ok=True)
        content = "\n".join(f'{str(item["name"]).strip()}: {" ".join(str(item["prompt"]).split())}' for item in entries) + "\n"
        targets.append((target, content))
    targets.append((root / "metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2) + "\n"))
    for target, content in targets:
        fd, temp_name = tempfile.mkstemp(prefix=f"{target.stem}-", suffix=".tmp", dir=target.parent)
        try:
            with os.fdopen(fd, "w", encoding="utf-8", newline="\n") as handle:
                handle.write(content)
            os.replace(temp_name, target)
        except Exception:
            try:
                os.unlink(temp_name)
            except OSError:
                pass
            raise


def _read_state() -> dict:
    path = _user_root() / "state.json"
    if not path.is_file():
        return {"version": 5, "favorites": [], "history": [], "overrides": {}, "library_order": []}
    try:
        with path.open("r", encoding="utf-8") as handle:
            state = json.load(handle)
    except (OSError, json.JSONDecodeError):
        return {"version": 5, "favorites": [], "history": [], "overrides": {}, "library_order": []}
    version = int(state.get("version", 1) or 1)
    return {
        "version": 5,
        "favorites": list(dict.fromkeys(str(x) for x in state.get("favorites", []) if x)),
        "history": [str(x) for x in state.get("history", []) if x][:100] if version >= 3 else [],
        "overrides": state.get("overrides", {}) if isinstance(state.get("overrides", {}), dict) else {},
        "library_order": list(dict.fromkeys(str(x) for x in state.get("library_order", []) if x)),
    }


def _write_state(state: dict) -> None:
    root = _user_root()
    root.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix="state-", suffix=".tmp", dir=root)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="\n") as handle:
            json.dump(state, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
        os.replace(temp_name, root / "state.json")
    except Exception:
        try:
            os.unlink(temp_name)
        except OSError:
            pass
        raise


def _with_override(item: dict, state: dict) -> dict:
    override = state["overrides"].get(item["name"], {})
    if not isinstance(override, dict) or not override:
        return item
    result = dict(item)
    result["name_zh"] = str(override.get("title") or result.get("name_zh") or result["name"])
    result["prompt"] = str(override.get("prompt") or result["prompt"])
    result["override_preview"] = str(override.get("preview") or "")
    return result


def _clean_text(value, field: str, max_length: int) -> str:
    text = str(value or "").strip()
    if not text:
        raise ValueError(f"{field} cannot be empty")
    if len(text) > max_length:
        raise ValueError(f"{field} is too long")
    return text


def _safe_library(value) -> str:
    text = str(value or "custom").strip().replace("\\", "/").strip("/")
    if not text or not re.fullmatch(r"[\w .+/-]{1,160}", text, re.UNICODE) or ".." in text.split("/"):
        raise ValueError("Invalid library name")
    return text


def _validated_style(data: dict, *, style_id: str | None = None, preview: str = "") -> dict:
    if not isinstance(data, dict):
        raise ValueError("Invalid style data")
    category = _clean_text(data.get("category"), "category", 40)
    if any(character in category for character in "\r\n:"):
        raise ValueError("Invalid category")
    safe_id = style_id or uuid4().hex
    if not re.fullmatch(r"[A-Za-z0-9_-]{1,80}", safe_id):
        raise ValueError("Invalid style id")
    return {
        "id": safe_id,
        "name": _clean_text(data.get("name"), "English name", 160),
        "name_zh": _clean_text(data.get("name_zh"), "Chinese name", 160),
        "category": category,
        "prompt": _clean_text(data.get("prompt"), "Prompt", 20000),
        "preview": preview,
        "library": _safe_library(data.get("library") or "custom"),
    }


def _user_styles() -> tuple[dict, ...]:
    valid = []
    for raw in _read_user_payload()["styles"]:
        try:
            valid.append(_validated_style(raw, style_id=str(raw.get("id") or uuid4().hex), preview=str(raw.get("preview") or "")))
        except ValueError:
            continue
    return tuple(valid)


def _rename_user_library(old_name, new_name) -> str:
    old_name = _safe_library(old_name)
    new_name = _safe_library(new_name)
    if old_name in _CATEGORIES or old_name in _VIRTUAL_LIBRARIES:
        raise ValueError("This library cannot be renamed")
    payload = _read_user_payload()
    if not any(item.get("library") == old_name for item in payload["styles"]):
        raise ValueError("Library not found")
    if new_name in _CATEGORIES or new_name in _VIRTUAL_LIBRARIES or any(
        item.get("library") == new_name for item in payload["styles"]
    ):
        raise ValueError("Library name already exists")
    for item in payload["styles"]:
        if item.get("library") == old_name:
            item["library"] = new_name
    _write_user_payload(payload)
    state = _read_state()
    state["library_order"] = [new_name if name == old_name else name for name in state.get("library_order", [])]
    _write_state(state)
    try:
        (_user_root() / "wildcards" / f"{old_name}.txt").unlink()
    except OSError:
        pass
    return new_name


def _delete_user_library(library) -> int:
    library = _safe_library(library)
    if library in _CATEGORIES or library in _VIRTUAL_LIBRARIES:
        raise ValueError("This library cannot be deleted")
    payload = _read_user_payload()
    removed = [item for item in payload["styles"] if item.get("library") == library]
    if not removed:
        raise ValueError("Library not found")
    payload["styles"] = [item for item in payload["styles"] if item.get("library") != library]
    _write_user_payload(payload)
    removed_names = {item["name"] for item in removed}
    state = _read_state()
    state["library_order"] = [name for name in state.get("library_order", []) if name != library]
    state["favorites"] = [name for name in state["favorites"] if name not in removed_names]
    state["history"] = [name for name in state["history"] if name not in removed_names]
    for name in removed_names:
        state["overrides"].pop(name, None)
    _write_state(state)
    try:
        (_user_root() / "wildcards" / f"{library}.txt").unlink()
    except OSError:
        pass
    for item in removed:
        if item.get("preview"):
            try:
                (_user_root() / "previews" / item["preview"]).unlink()
            except OSError:
                pass
    return len(removed)


def _ordered_user_libraries(styles: tuple[dict, ...] | list[dict], state: dict | None = None) -> list[str]:
    state = state or _read_state()
    available = sorted({str(item.get("library") or "") for item in styles if item.get("library")})
    ordered = [library for library in state.get("library_order", []) if library in available]
    return ordered + [library for library in available if library not in ordered]


def _apply_library_manager(changes) -> list[str]:
    if not isinstance(changes, list):
        raise ValueError("Invalid library settings")
    payload = _read_user_payload()
    existing = _ordered_user_libraries(payload["styles"])
    by_original = {}
    for change in changes:
        if not isinstance(change, dict):
            raise ValueError("Invalid library settings")
        original = _safe_library(change.get("library"))
        if original in by_original:
            raise ValueError("Duplicate library settings")
        by_original[original] = change
    if set(by_original) != set(existing):
        raise ValueError("Library list changed; reload and try again")

    requested_order = list(by_original)
    kept = []
    final_names = set()
    for original in requested_order:
        change = by_original[original]
        if bool(change.get("delete")):
            continue
        name = _safe_library(change.get("name") or original)
        if name in _VIRTUAL_LIBRARIES or name in _CATEGORIES or name in final_names:
            raise ValueError("Library name already exists")
        final_names.add(name)
        kept.append((original, name))

    name_map = {original: name for original, name in kept}
    removed = {original for original in existing if original not in name_map}
    removed_items = [item for item in payload["styles"] if item.get("library") in removed]
    payload["styles"] = [item for item in payload["styles"] if item.get("library") not in removed]
    for item in payload["styles"]:
        if item.get("library") in name_map:
            item["library"] = name_map[item["library"]]
    _write_user_payload(payload)

    state = _read_state()
    removed_names = {item["name"] for item in removed_items}
    state["favorites"] = [name for name in state["favorites"] if name not in removed_names]
    state["history"] = [name for name in state["history"] if name not in removed_names]
    for name in removed_names:
        state["overrides"].pop(name, None)
    state["library_order"] = [name for _original, name in kept]
    _write_state(state)

    for old_name in existing:
        if old_name in removed or name_map.get(old_name) != old_name:
            try:
                (_user_root() / "wildcards" / f"{old_name}.txt").unlink()
            except OSError:
                pass
    for item in removed_items:
        if item.get("preview"):
            try:
                (_user_root() / "previews" / item["preview"]).unlink()
            except OSError:
                pass
    return state["library_order"]


def _all_styles() -> tuple[dict, ...]:
    state = _read_state()
    return tuple(_with_override(item, state) for item in _user_styles())


def _style_by_name() -> dict[str, dict]:
    return {item["name"]: item for item in _all_styles()}


def _preview_path(item: dict) -> Path:
    root = _user_root() / "previews"
    filename = item.get("override_preview") or item.get("preview", "")
    path = (root / str(filename)).resolve()
    if path.parent != root.resolve():
        raise ValueError("Invalid preview path")
    return path


def _catalog_item(item: dict, source: str) -> dict:
    enriched = dict(item)
    enriched["source"] = source
    preview_path = _preview_path(enriched) if enriched.get("preview") or enriched.get("override_preview") else None
    return {
        "id": enriched.get("id"),
        "name": enriched["name"],
        "name_zh": enriched["name_zh"],
        "category": enriched["category"],
        "search_text": enriched["prompt"],
        "preview": enriched.get("preview", ""),
        "has_preview": bool(preview_path) and preview_path.is_file(),
        "preview_version": preview_path.stat().st_mtime_ns if preview_path and preview_path.is_file() else 0,
        "source": source,
        "library": enriched.get("library") or (enriched["category"] if source == "builtin" else "custom"),
    }


def _formatted_prompt(item: dict) -> str:
    return item["prompt"]


def _items_for_library(library: str, search_query: str = "") -> list[dict]:
    styles = list(_all_styles())
    state = _read_state()
    by_name = {item["name"]: item for item in styles}
    if library == "收藏夹":
        items = [by_name[name] for name in state["favorites"] if name in by_name]
    elif library == "历史记录":
        items = [by_name[name] for name in state["history"] if name in by_name]
    else:
        items = [item for item in styles if (item.get("library") or item.get("category")) == library]
    query = unicodedata.normalize("NFKC", str(search_query or "")).strip().casefold()
    if not query:
        return items
    return [
        item for item in items
        if any(query in unicodedata.normalize("NFKC", str(item.get(field) or "")).casefold()
               for field in ("name", "name_zh", "prompt"))
    ]


def _record_history(name: str) -> None:
    state = _read_state()
    state["history"] = [name] + [item for item in state["history"] if item != name]
    state["history"] = state["history"][:100]
    _write_state(state)


_STYLES = ()  # Backward-compatible extension surface; libraries are imported on demand.


class NO8DKreaStyleSelector:
    @classmethod
    def INPUT_TYPES(cls):
        return {"required": {
            "style": ("STRING", {"default": ""}),
            "library": ("STRING", {"default": ""}),
            "random_mode": ("BOOLEAN", {"default": False}),
            "selected_styles": ("STRING", {"default": "[]"}),
            # Keeps an intentionally cleared selection distinct from a newly-created node.
            "selection_cleared": ("BOOLEAN", {"default": False}),
            "output_all": ("BOOLEAN", {"default": False}),
            "search_query": ("STRING", {"default": ""}),
        }}

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("prompt",)
    OUTPUT_IS_LIST = (True,)
    FUNCTION = "select_style"
    CATEGORY = "NO8D-control"
    DESCRIPTION = "Browse prompt libraries visually and output the selected English prompt."

    @classmethod
    def VALIDATE_INPUTS(cls, style, library, random_mode, selected_styles="[]", selection_cleared=False, output_all=False, search_query=""):
        try:
            selected = json.loads(selected_styles or "[]")
            if not isinstance(selected, list):
                return "Selected styles must be a JSON list"
        except json.JSONDecodeError:
            return "Selected styles must be valid JSON"
        if not style:
            return True
        if style not in _style_by_name():
            return f"Unknown style: {style}"
        return True

    @classmethod
    def IS_CHANGED(cls, style, library, random_mode, selected_styles="[]", selection_cleared=False, output_all=False, search_query=""):
        try:
            selected = json.loads(selected_styles or "[]")
        except json.JSONDecodeError:
            selected = []
        if isinstance(selected, list) and selected:
            return selected_styles
        if output_all:
            return float("nan")
        if random_mode:
            return float("nan")
        return "" if selection_cleared else style

    def select_style(self, style, library, random_mode, selected_styles="[]", selection_cleared=False, output_all=False, search_query=""):
        by_name = _style_by_name()
        try:
            selected = json.loads(selected_styles or "[]")
        except json.JSONDecodeError:
            selected = []
        selected = list(dict.fromkeys(name for name in selected if isinstance(name, str) and name in by_name))
        if selected:
            items = [by_name[name] for name in selected]
            for item in items:
                _record_history(item["name"])
            return ([_formatted_prompt(item) for item in items],)
        if output_all:
            items = _items_for_library(library, search_query)
            for item in items:
                _record_history(item["name"])
            return ([_formatted_prompt(item) for item in items],)
        if random_mode:
            candidates = _items_for_library(library, search_query)
            if candidates:
                item = random.choice(candidates)
                _record_history(item["name"])
                return ([_formatted_prompt(item)],)
            return ([""],)
        if selection_cleared:
            return ([""],)
        if not style and not by_name:
            return ([],)
        item = by_name.get(style)
        if item is None:
            return ([""],)
        _record_history(item["name"])
        return ([_formatted_prompt(item)],)


async def _read_style_form(request) -> tuple[dict, bytes | None, str | None]:
    reader = await request.multipart()
    data = None
    image_bytes = None
    image_type = None
    while True:
        part = await reader.next()
        if part is None:
            break
        if part.name == "data":
            data = json.loads(await part.text())
        elif part.name == "preview":
            image_type = part.headers.get("Content-Type", "").split(";", 1)[0].lower()
            chunks = []
            total = 0
            while True:
                chunk = await part.read_chunk()
                if not chunk:
                    break
                total += len(chunk)
                if total > _MAX_PREVIEW_BYTES:
                    raise ValueError("Preview image is too large")
                chunks.append(chunk)
            image_bytes = b"".join(chunks)
    if data is None:
        raise ValueError("Missing style data")
    return data, image_bytes, image_type


def _save_preview(style_id: str, content: bytes, content_type: str | None) -> str:
    suffix = _ALLOWED_IMAGE_TYPES.get(content_type or "")
    if not suffix or not content:
        raise ValueError("Preview must be a PNG, JPEG, or WebP image")
    signatures = {
        ".png": content.startswith(b"\x89PNG\r\n\x1a\n"),
        ".jpg": content.startswith(b"\xff\xd8\xff"),
        ".webp": len(content) >= 12 and content[:4] == b"RIFF" and content[8:12] == b"WEBP",
    }
    if not signatures[suffix]:
        raise ValueError("Preview file content does not match its image type")
    preview_dir = _user_root() / "previews"
    preview_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{style_id}{suffix}"
    path = preview_dir / filename
    path.write_bytes(content)
    return filename


def _rows_from_values(values: list[list[object]]) -> list[dict]:
    rows = [[str(cell or "").strip() for cell in row] for row in values if any(str(cell or "").strip() for cell in row)]
    if not rows:
        return []
    aliases = {"name": {"name", "style", "style name", "title", "名称", "英文名称"}, "name_zh": {"name_zh", "chinese name", "中文名", "中文名称"}, "category": {"category", "分类", "大分类"}, "prompt": {"prompt", "text", "提示词", "内容"}}
    normalized = [cell.casefold() for cell in rows[0]]
    columns = {key: next((i for i, cell in enumerate(normalized) if cell in names), None) for key, names in aliases.items()}
    has_header = columns["prompt"] is not None or columns["name"] is not None
    data_rows = rows[1:] if has_header else rows
    result = []
    for index, row in enumerate(data_rows, 1):
        def cell(key):
            position = columns[key]
            return row[position].strip() if position is not None and position < len(row) else ""
        if has_header:
            prompt = cell("prompt") or (row[0] if len(row) == 1 else "")
            name = cell("name")
        else:
            raw = row[0] if row else ""
            name, separator, prompt = raw.partition(":")
            if not separator:
                name, prompt = f"Entry {index:03d}", raw
        if not prompt.strip():
            continue
        name = name.strip() or f"Entry {index:03d}"
        category = cell("category") if has_header else ""
        result.append({"name": name, "name_zh": cell("name_zh") if has_header else name, "category": category or _DEFAULT_USER_CATEGORY, "prompt": prompt.strip()})
    totals = {}
    for item in result:
        key = item["name"].casefold()
        totals[key] = totals.get(key, 0) + 1
    seen = {}
    for item in result:
        key = item["name"].casefold()
        if totals[key] > 1:
            seen[key] = seen.get(key, 0) + 1
            item["name"] = f'{item["name"]} {seen[key]:03d}'
            if not item.get("name_zh") or item["name_zh"].casefold() == key:
                item["name_zh"] = item["name"]
    return result


def _parse_import(content: bytes, filename: str) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".txt", ".csv"}:
        text = content.decode("utf-8-sig", errors="replace")
        values = [[line] for line in text.splitlines()] if suffix == ".txt" else list(csv.reader(io.StringIO(text)))
        return _rows_from_values(values)
    if suffix == ".xlsx":
        openpyxl = _require_openpyxl()
        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        return _rows_from_values([list(row) for row in sheet.iter_rows(values_only=True)])
    raise ValueError("Only TXT, CSV, and XLSX files are supported")


def _parse_import_bundle(content: bytes, filename: str) -> tuple[list[dict], dict[int, tuple[bytes, str]], dict]:
    rows = _parse_import(content, filename)
    if Path(filename).suffix.lower() != ".xlsx":
        return rows, {}, {}
    openpyxl = _require_openpyxl()
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    sheet = next((candidate for candidate in workbook.worksheets if candidate.title != "_NO8D_DATA"), workbook.active)
    previews = {}
    for image in getattr(sheet, "_images", []):
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is None:
            continue
        row_index = int(marker.row) - 1
        if row_index < 0 or row_index >= len(rows):
            continue
        data = image._data()
        content_type = "image/png" if data.startswith(b"\x89PNG") else "image/jpeg" if data.startswith(b"\xff\xd8\xff") else "image/webp"
        previews[row_index] = (data, content_type)
    portable = {"favorites": [], "history": [], "libraries": {}}
    if "_NO8D_DATA" in workbook.sheetnames:
        metadata = workbook["_NO8D_DATA"]
        values = list(metadata.iter_rows(values_only=True))
        if values:
            headers = {str(value or "").strip(): index for index, value in enumerate(values[0])}
            for row in values[1:]:
                name = str(row[headers.get("name", -1)] or "").strip() if headers.get("name", -1) >= 0 else ""
                if name and headers.get("favorite", -1) >= 0 and bool(row[headers["favorite"]]):
                    portable["favorites"].append(name)
                if name and headers.get("history_index", -1) >= 0 and row[headers["history_index"]] not in (None, ""):
                    portable["history"].append((int(row[headers["history_index"]]), name))
                if name and headers.get("source_library", -1) >= 0 and row[headers["source_library"]]:
                    portable["libraries"][name] = str(row[headers["source_library"]]).strip()
        portable["history"] = [name for _, name in sorted(portable["history"])]
    return rows, previews, portable


def _export_items_xlsx(items: list[dict], library: str) -> bytes:
    openpyxl = _require_openpyxl()
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from PIL import Image as PILImage

    if not items:
        raise ValueError("No styles selected")
    state = _read_state()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = re.sub(r"[\\/*?:\[\]]", "_", library)[:31] or "Library"
    sheet.freeze_panes = "A2"
    headers = ["Preview", "Style Name", "Chinese Name", "Prompt", "Image File"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 30
    sheet.column_dimensions["D"].width = 90
    sheet.column_dimensions["E"].width = 34
    buffers = []
    for row_number, item in enumerate(items, start=2):
        sheet.append(["", item["name"], item.get("name_zh") or item["name"], item["prompt"], item.get("preview") or ""])
        sheet.row_dimensions[row_number].height = 78
        sheet.cell(row_number, 4).alignment = Alignment(wrap_text=True, vertical="center")
        path = _preview_path({**item, "source": "user"}) if item.get("preview") else None
        if path and path.is_file():
            with PILImage.open(path) as image:
                buffer = BytesIO()
                image.convert("RGB").resize((256, 256)).save(buffer, format="PNG")
                buffer.seek(0)
            buffers.append(buffer)
            picture = XLImage(buffer)
            picture.width = 96
            picture.height = 96
            sheet.add_image(picture, f"A{row_number}")
    metadata = workbook.create_sheet("_NO8D_DATA")
    metadata.append(["version", "library", "name", "favorite", "history_index", "source_library"])
    history_index = {name: index for index, name in enumerate(state["history"])}
    favorites = set(state["favorites"])
    for item in items:
        source_library = item.get("library") or library
        metadata.append([1, library, item["name"], item["name"] in favorites, history_index.get(item["name"]), source_library])
    metadata.sheet_state = "hidden"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _export_library_xlsx(library: str) -> bytes:
    items = [item for item in _all_styles() if item.get("library") == library]
    if not items:
        raise ValueError("Library not found")
    return _export_items_xlsx(items, library)


def _import_favorite_xlsx(content: bytes, filename: str) -> dict:
    if Path(filename).suffix.lower() != ".xlsx":
        raise ValueError("Favorites must be imported from an XLSX file")
    rows, previews, portable = _parse_import_bundle(content, filename)
    if not rows:
        raise ValueError("No favorite cards found")
    payload = _read_user_payload()
    available = _style_by_name()
    restored = []
    imported = 0
    for row_index, raw in enumerate(rows):
        candidate = dict(raw)
        name = str(candidate.get("name") or "").strip()
        if name in available:
            restored.append(name)
            continue
        source_library = portable.get("libraries", {}).get(name) or "favorites_imported"
        if source_library in _VIRTUAL_LIBRARIES:
            source_library = "favorites_imported"
        candidate["library"] = _safe_library(source_library)
        candidate["category"] = candidate.get("category") or _DEFAULT_USER_CATEGORY
        candidate["name_zh"] = candidate.get("name_zh") or name
        style = _validated_style(candidate)
        if row_index in previews:
            image_bytes, image_type = previews[row_index]
            style["preview"] = _save_preview(style["id"], image_bytes, image_type)
        payload["styles"].append(style)
        available[style["name"]] = style
        restored.append(style["name"])
        imported += 1
    if imported:
        _write_user_payload(payload)
    state = _read_state()
    state["favorites"] = list(dict.fromkeys(restored + state["favorites"]))
    _write_state(state)
    return {"ok": True, "imported": imported, "favorites": len(restored)}


def _google_sheet_csv_url(value: str) -> str:
    parsed = urlparse(value.strip())
    if parsed.scheme != "https" or parsed.hostname != "docs.google.com":
        raise ValueError("Only public docs.google.com spreadsheet links are allowed")
    match = re.search(r"/spreadsheets/d/([A-Za-z0-9_-]+)", parsed.path)
    if not match:
        raise ValueError("Invalid Google Sheets URL")
    query = parse_qs(parsed.query)
    fragment_query = parse_qs(parsed.fragment)
    gid = (query.get("gid") or fragment_query.get("gid") or ["0"])[0]
    if not gid.isdigit():
        gid = "0"
    return f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=csv&gid={gid}"


if PromptServer is not None and web is not None:
    @PromptServer.instance.routes.get(f"{_API_PREFIX}/styles")
    async def no8d_krea_style_catalog(_request):
        state = _read_state()
        custom = [_catalog_item(_with_override(item, state), "user") for item in _user_styles()]
        styles = custom
        favorites = set(state["favorites"])
        history_order = {name: index for index, name in enumerate(state["history"])}
        for item in styles:
            item["favorite"] = item["name"] in favorites
            item["history_index"] = history_order.get(item["name"])
        libraries = list(_VIRTUAL_LIBRARIES) + _ordered_user_libraries(custom, state)
        return web.json_response({"builtin_libraries": [], "virtual_libraries": list(_VIRTUAL_LIBRARIES), "libraries": libraries, "styles": styles})

    @PromptServer.instance.routes.post(f"{_API_PREFIX}/favorite")
    async def no8d_krea_favorite(request):
        try:
            body = await request.json()
            available = _style_by_name()
            requested = body.get("names")
            if isinstance(requested, list):
                names = list(dict.fromkeys(str(name) for name in requested if name))
            else:
                name = str(body.get("name") or "")
                names = [name] if name else []
            if not names or any(name not in available for name in names):
                raise ValueError("Style not found")
            state = _read_state()
            favorites = set(state["favorites"])
            if isinstance(body.get("favorite"), bool):
                favorite = body["favorite"]
            else:
                favorite = names[0] not in favorites
            if favorite:
                favorites.update(names)
            else:
                favorites.difference_update(names)
            state["favorites"] = [name for name in state["favorites"] if name in favorites]
            state["favorites"].extend(name for name in names if name in favorites and name not in state["favorites"])
            _write_state(state)
            return web.json_response({"ok": True, "favorite": favorite, "count": len(names)})
        except (ValueError, json.JSONDecodeError) as error:
            return web.json_response({"error": str(error)}, status=400)

    @PromptServer.instance.routes.post(f"{_API_PREFIX}/history/remove")
    async def no8d_krea_history_remove(request):
        try:
            body = await request.json()
            names = list(dict.fromkeys(str(name) for name in body.get("names", []) if name))
            if not names:
                raise ValueError("Select at least one history card")
            state = _read_state()
            state["history"] = [name for name in state["history"] if name not in names]
            _write_state(state)
            return web.json_response({"ok": True, "removed": len(names)})
        except (ValueError, json.JSONDecodeError) as error:
            return web.json_response({"error": str(error)}, status=400)

    @PromptServer.instance.routes.get(f"{_API_PREFIX}/preview")
    async def no8d_krea_style_preview(request):
        item = _style_by_name().get(request.rel_url.query.get("style", ""))
        if item is None:
            return web.Response(status=404)
        source = "user" if item.get("id") else "builtin"
        try:
            path = _preview_path({**item, "source": source})
        except ValueError:
            return web.Response(status=404)
        return web.FileResponse(path) if path.is_file() else web.Response(status=404)

    @PromptServer.instance.routes.get(f"{_API_PREFIX}/library/export")
    async def no8d_krea_library_export(request):
        try:
            library = _safe_library(request.rel_url.query.get("library", ""))
            if library in _VIRTUAL_LIBRARIES:
                raise ValueError("Virtual libraries cannot be exported")
            content = _export_library_xlsx(library)
            filename = f"{Path(library).name}.xlsx"
            disposition = "inline" if request.rel_url.query.get("view") == "1" else "attachment"
            return web.Response(body=content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"{disposition}; filename*=UTF-8''{quote(filename)}"})
        except (ValueError, OSError):
            return web.Response(status=400)

    @PromptServer.instance.routes.post(f"{_API_PREFIX}/libraries/export")
    async def no8d_krea_libraries_export(request):
        try:
            body = await request.json()
            libraries = list(dict.fromkeys(_safe_library(value) for value in body.get("libraries", [])))
            if not libraries or any(library in _VIRTUAL_LIBRARIES for library in libraries):
                raise ValueError("Select at least one library")
            output = BytesIO()
            with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                for library in libraries:
                    archive.writestr(f"{Path(library).name}.xlsx", _export_library_xlsx(library))
            return web.Response(body=output.getvalue(), content_type="application/zip", headers={"Content-Disposition": "attachment; filename=no8d-libraries.zip"})
        except (ValueError, json.JSONDecodeError, OSError):
            return web.Response(status=400)

    @PromptServer.instance.routes.get(f"{_API_PREFIX}/favorites/export")
    async def no8d_krea_favorites_export(_request):
        try:
            content = _export_items_xlsx(_items_for_library("收藏夹"), "favorites")
            return web.Response(
                body=content,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=no8d-favorites.xlsx"},
            )
        except (ValueError, OSError):
            return web.Response(status=400)

    @PromptServer.instance.routes.post(f"{_API_PREFIX}/favorites/import")
    async def no8d_krea_favorites_import(request):
        try:
            reader = await request.multipart()
            part = await reader.next()
            if part is None or part.name != "file":
                raise ValueError("Missing favorites file")
            content = await part.read(decode=False)
            filename = part.filename or "favorites.xlsx"
            return web.json_response(_import_favorite_xlsx(content, filename))
        except (ValueError, UnicodeDecodeError, json.JSONDecodeError, OSError) as error:
            return web.json_response({"error": str(error)}, status=400)

    @PromptServer.instance.routes.post(f"{_API_PREFIX}/libraries/manage")
    async def no8d_krea_libraries_manage(request):
        try:
            body = await request.json()
            order = _apply_library_manager(body.get("libraries"))
            return web.json_response({"ok": True, "libraries": order})
        except (ValueError, json.JSONDecodeError) as error:
            return web.json_response({"error": str(error)}, status=400)

    @PromptServer.instance.routes.post(f"{_API_PREFIX}/styles/export")
    async def no8d_krea_styles_export(request):
        try:
            body = await request.json()
            names = list(dict.fromkeys(str(value) for value in body.get("names", []) if value))
            available = _style_by_name()
            if not names or any(name not in available for name in names):
                raise ValueError("Select at least one valid style")
            items = [available[name] for name in names]
            libraries = list(dict.fromkeys(item.get("library") or "styles" for item in items))
            sheet_name = libraries[0] if len(libraries) == 1 else "selected_styles"
            content = _export_items_xlsx(items, sheet_name)
            filename = f"{names[0]}.xlsx" if len(names) == 1 else "selected-styles.xlsx"
            return web.Response(
                body=content,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
            )
        except (ValueError, json.JSONDecodeError, OSError):
            return web.Response(status=400)

    @PromptServer.instance.routes.post(f"{_API_PREFIX}/custom")
    async def no8d_krea_style_save(request):
        try:
            data, image_bytes, image_type = await _read_style_form(request)
            payload = _read_user_payload()
            requested_id = str(data.get("id") or "").strip()
            index = next((i for i, item in enumerate(payload["styles"]) if item.get("id") == requested_id), -1)
            old = payload["styles"][index] if index >= 0 else None
            style_id = requested_id if old else uuid4().hex
            preview = str(old.get("preview") or "") if old else ""
            style = _validated_style(data, style_id=style_id, preview=preview)
            duplicate = next((item for item in _all_styles() if item["name"].casefold() == style["name"].casefold() and item.get("id") != style_id), None)
            if duplicate:
                if old:
                    raise ValueError("English style name already exists")
                # The editor only exposes the localized card title.  New cards
                # therefore receive a deterministic internal suffix instead of
                # failing when that title was used before.
                used_names = {item["name"].casefold() for item in _all_styles()}
                base_name = style["name"]
                base_title = style["name_zh"]
                suffix = 2
                while f"{base_name} {suffix}".casefold() in used_names:
                    suffix += 1
                candidate = dict(data)
                candidate["name"] = f"{base_name} {suffix}"
                candidate["name_zh"] = f"{base_title} {suffix}"
                style = _validated_style(candidate, style_id=style_id, preview=preview)
            if image_bytes is not None:
                style["preview"] = _save_preview(style_id, image_bytes, image_type)
                if old and old.get("preview") and old["preview"] != style["preview"]:
                    try:
                        (_user_root() / "previews" / old["preview"]).unlink()
                    except OSError:
                        pass
            if index >= 0:
                payload["styles"][index] = style
            else:
                after_id = str(data.get("insert_after_id") or "").strip()
                after_index = next((i for i, item in enumerate(payload["styles"]) if item.get("id") == after_id), -1)
                if after_index >= 0 and payload["styles"][after_index].get("library") == style["library"]:
                    payload["styles"].insert(after_index + 1, style)
                else:
                    payload["styles"].append(style)
            _write_user_payload(payload)
            if old and old.get("library") != style["library"] and not any(item.get("library") == old.get("library") for item in payload["styles"]):
                try:
                    (_user_root() / "wildcards" / f'{_safe_library(old.get("library"))}.txt').unlink()
                except OSError:
                    pass
            return web.json_response({"ok": True, "style": _catalog_item(style, "user")})
        except (ValueError, json.JSONDecodeError) as error:
            return web.json_response({"error": str(error)}, status=400)

    @PromptServer.instance.routes.post(f"{_API_PREFIX}/card/update")
    async def no8d_krea_card_update(request):
        try:
            data, image_bytes, image_type = await _read_style_form(request)
            original_name = _clean_text(data.get("original_name"), "Original name", 160)
            if original_name not in _style_by_name():
                raise ValueError("Style not found")
            state = _read_state()
            previous = state["overrides"].get(original_name, {})
            preview = str(previous.get("preview") or "") if isinstance(previous, dict) else ""
            if image_bytes is not None:
                new_preview = _save_preview(f"override-{uuid4().hex}", image_bytes, image_type)
                if preview and preview != new_preview:
                    try:
                        (_user_root() / "previews" / preview).unlink()
                    except OSError:
                        pass
                preview = new_preview
            state["overrides"][original_name] = {
                "title": _clean_text(data.get("title"), "Title", 160),
                "prompt": _clean_text(data.get("prompt"), "Prompt", 20000),
                "preview": preview,
            }
            _write_state(state)
            return web.json_response({"ok": True})
        except (ValueError, json.JSONDecodeError) as error:
            return web.json_response({"error": str(error)}, status=400)

    @PromptServer.instance.routes.post(f"{_API_PREFIX}/import/preview")
    async def no8d_krea_import_preview(request):
        try:
            if request.content_type == "application/json":
                body = await request.json()
                url = _google_sheet_csv_url(str(body.get("url") or ""))
                def download():
                    with urlopen(url, timeout=30) as response:
                        content = response.read(20 * 1024 * 1024 + 1)
                    if len(content) > 20 * 1024 * 1024:
                        raise ValueError("Spreadsheet is too large")
                    return content
                content = await asyncio.to_thread(download)
                filename = "google-sheet.csv"
            else:
                reader = await request.multipart()
                part = await reader.next()
                if part is None or part.name != "file":
                    raise ValueError("Missing import file")
                filename = part.filename or "import.txt"
                content = await part.read(decode=False)
                if len(content) > 20 * 1024 * 1024:
                    raise ValueError("Import file is too large")
            rows, _previews, portable = _parse_import_bundle(content, filename)
            if not rows:
                raise ValueError("No valid wildcard rows found")
            return web.json_response({"rows": rows, "count": len(rows)})
        except (ValueError, json.JSONDecodeError, OSError) as error:
            return web.json_response({"error": str(error)}, status=400)

    @PromptServer.instance.routes.post(f"{_API_PREFIX}/library/rename")
    async def no8d_krea_library_rename(request):
        try:
            body = await request.json()
            new_name = _rename_user_library(body.get("old_name"), body.get("new_name"))
            return web.json_response({"ok": True, "library": new_name})
        except (ValueError, json.JSONDecodeError) as error:
            return web.json_response({"error": str(error)}, status=400)

    @PromptServer.instance.routes.delete(f"{_API_PREFIX}/library/{{library}}")
    async def no8d_krea_library_delete(request):
        try:
            removed = _delete_user_library(request.match_info["library"])
            return web.json_response({"ok": True, "removed": removed})
        except ValueError as error:
            return web.json_response({"error": str(error)}, status=400)

    @PromptServer.instance.routes.post(f"{_API_PREFIX}/library/delete")
    async def no8d_krea_library_delete_post(request):
        try:
            body = await request.json()
            removed = _delete_user_library(body.get("library"))
            return web.json_response({"ok": True, "removed": removed})
        except (ValueError, json.JSONDecodeError) as error:
            return web.json_response({"error": str(error)}, status=400)

    @PromptServer.instance.routes.post(f"{_API_PREFIX}/import/commit")
    async def no8d_krea_import_commit(request):
        try:
            previews = {}
            portable = {}
            if request.content_type.startswith("multipart/"):
                reader = await request.multipart()
                fields = {}
                content = None
                filename = "import.xlsx"
                while True:
                    part = await reader.next()
                    if part is None:
                        break
                    if part.name == "file":
                        filename = part.filename or filename
                        content = await part.read(decode=False)
                        if len(content) > 50 * 1024 * 1024:
                            raise ValueError("Import file is too large")
                    else:
                        fields[part.name] = await part.text()
                if content is None:
                    raise ValueError("Missing import file")
                rows, previews, portable = _parse_import_bundle(content, filename)
                library = _safe_library(fields.get("library") or Path(filename).stem)
            else:
                body = await request.json()
                rows = body.get("rows")
                library = _safe_library(body.get("library") or "custom")
            if not isinstance(rows, list) or not rows or len(rows) > 5000:
                raise ValueError("Import must contain between 1 and 5000 rows")
            payload = _read_user_payload()
            existing = {item["name"].casefold() for item in _all_styles()}
            imported = []
            skipped = 0
            imported_names = set()
            for row_index, raw in enumerate(rows):
                candidate = dict(raw) if isinstance(raw, dict) else {}
                candidate["library"] = library
                candidate["category"] = candidate.get("category") or _DEFAULT_USER_CATEGORY
                candidate["name_zh"] = candidate.get("name_zh") or candidate.get("name")
                try:
                    style = _validated_style(candidate)
                except ValueError:
                    skipped += 1
                    continue
                if style["name"].casefold() in existing:
                    skipped += 1
                    continue
                existing.add(style["name"].casefold())
                if row_index in previews:
                    image_bytes, image_type = previews[row_index]
                    style["preview"] = _save_preview(style["id"], image_bytes, image_type)
                payload["styles"].append(style)
                imported.append(style)
                imported_names.add(style["name"])
            if not imported:
                raise ValueError("All rows were invalid or duplicated")
            _write_user_payload(payload)
            state = _read_state()
            for name in portable.get("favorites", []):
                if name in imported_names and name not in state["favorites"]:
                    state["favorites"].append(name)
            restored_history = [name for name in portable.get("history", []) if name in imported_names]
            state["history"] = restored_history + [name for name in state["history"] if name not in restored_history]
            state["history"] = state["history"][:100]
            _write_state(state)
            return web.json_response({"ok": True, "imported": len(imported), "skipped": skipped, "first": imported[0]["name"]})
        except (ValueError, json.JSONDecodeError) as error:
            return web.json_response({"error": str(error)}, status=400)

    @PromptServer.instance.routes.delete(f"{_API_PREFIX}/custom/{{style_id}}")
    async def no8d_krea_style_delete(request):
        style_id = request.match_info["style_id"]
        if not re.fullmatch(r"[A-Za-z0-9_-]{1,80}", style_id):
            return web.json_response({"error": "Invalid style id"}, status=400)
        payload = _read_user_payload()
        removed = next((item for item in payload["styles"] if item.get("id") == style_id), None)
        if removed is None:
            return web.json_response({"error": "Style not found"}, status=404)
        payload["styles"] = [item for item in payload["styles"] if item.get("id") != style_id]
        _write_user_payload(payload)
        if not any(item.get("library") == removed.get("library") for item in payload["styles"]):
            try:
                (_user_root() / "wildcards" / f'{_safe_library(removed.get("library"))}.txt').unlink()
            except OSError:
                pass
        if removed.get("preview"):
            try:
                (_user_root() / "previews" / removed["preview"]).unlink()
            except OSError:
                pass
        return web.json_response({"ok": True})


NODE_CLASS_MAPPINGS = {"NO8DKreaStyleSelector": NO8DKreaStyleSelector}
NODE_DISPLAY_NAME_MAPPINGS = {"NO8DKreaStyleSelector": "NO8D-Prompt-libraries"}
